"""Read, validate, and prepare company-location Excel workbooks."""

from pathlib import Path
from typing import Iterator, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from models import Location


PathLike = Union[str, Path]


class ExcelServiceError(Exception):
    """Base exception for workbook processing failures."""


class WorkbookLoadError(ExcelServiceError):
    """Raised when an Excel workbook cannot be opened."""


class CompanyColumnNotFoundError(ExcelServiceError):
    """Raised when the active worksheet has no Company column."""


class WorkbookSaveError(ExcelServiceError):
    """Raised when an Excel workbook cannot be saved."""


class WorkbookNotFoundError(ExcelServiceError):
    """Raised when the configured input directory has no workbooks."""


class ExcelService:
    """Provide operations for the application's input Excel workbook."""

    def load_workbook(self, workbook_path: PathLike) -> Workbook:
        """Open and return the workbook at ``workbook_path``.

        Raises:
            WorkbookLoadError: If the path is missing or the workbook is invalid.
        """
        path = Path(workbook_path)

        if not path.is_file():
            raise WorkbookLoadError(f"Workbook not found: {path}")

        try:
            return load_workbook(path)
        except (InvalidFileException, OSError, ValueError) as error:
            raise WorkbookLoadError(
                f"Unable to open workbook '{path}': {error}"
            ) from error

    def get_active_worksheet(self, workbook: Workbook) -> Worksheet:
        """Return the workbook's active worksheet.

        Raises:
            ExcelServiceError: If the workbook has no active worksheet.
        """
        try:
            return workbook.active
        except (AttributeError, IndexError) as error:
            raise ExcelServiceError(
                "The workbook does not contain an active worksheet."
            ) from error

    def find_company_column(self, worksheet: Worksheet) -> int:
        """Return the one-based index of the Company header in row one.

        Header matching ignores leading and trailing whitespace and case.

        Raises:
            CompanyColumnNotFoundError: If the Company column is absent.
        """
        for column_index, cell in enumerate(worksheet[1], start=1):
            value = cell.value
            if isinstance(value, str) and value.strip().casefold() == "company":
                return column_index

        raise CompanyColumnNotFoundError(
            "The active worksheet must contain a 'Company' header in row 1."
        )

    def ensure_location_column(self, worksheet: Worksheet) -> int:
        """Return the Location column index, creating its header if needed."""
        for column_index, cell in enumerate(worksheet[1], start=1):
            value = cell.value
            if isinstance(value, str) and value.strip().casefold() == "location":
                return column_index

        location_column = worksheet.max_column + 1
        worksheet.cell(row=1, column=location_column, value="Location")
        return location_column

    def find_workbooks(self, input_directory: PathLike) -> tuple[Path, ...]:
        """Return supported input workbooks in deterministic filename order.

        Args:
            input_directory: Directory containing the source workbooks.

        Raises:
            WorkbookNotFoundError: If the directory is invalid or has no Excel files.
        """

        directory = Path(input_directory)
        if not directory.is_dir():
            raise WorkbookNotFoundError(f"Input directory not found: {directory}")

        workbooks = tuple(
            path
            for path in sorted(
                directory.iterdir(), key=lambda item: item.name.casefold()
            )
            if path.is_file()
            and path.suffix.casefold() in {".xlsx", ".xlsm"}
            and not path.name.startswith("~$")
        )
        if not workbooks:
            raise WorkbookNotFoundError(
                f"No Excel workbooks found in input directory: {directory}"
            )
        return workbooks

    def iter_company_rows(
        self, worksheet: Worksheet, company_column: int
    ) -> Iterator[tuple[int, str]]:
        """Yield non-empty company names and their row numbers, excluding headers."""

        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=company_column).value
            if isinstance(value, str) and value.strip():
                yield row_index, value.strip()

    def write_location(
        self, worksheet: Worksheet, row_index: int, location_column: int,
        location: Location,
    ) -> None:
        """Write a normalized location into a worksheet cell.

        Args:
            worksheet: Worksheet receiving the location.
            row_index: One-based row index for the company.
            location_column: One-based column index for the Location field.
            location: Normalized location to serialize for Excel.
        """

        worksheet.cell(
            row=row_index,
            column=location_column,
            value=f"{location.city}, {location.state}, {location.country}",
        )

    def output_path(self, input_path: PathLike, output_directory: PathLike) -> Path:
        """Return an output path and create its parent directory when needed."""

        source_path = Path(input_path)
        destination_directory = Path(output_directory)
        destination_directory.mkdir(parents=True, exist_ok=True)
        output_name = f"{source_path.stem}_enriched{source_path.suffix}"
        return destination_directory / output_name

    def save_workbook(self, workbook: Workbook, workbook_path: PathLike) -> None:
        """Save ``workbook`` to ``workbook_path``.

        Raises:
            WorkbookSaveError: If the workbook cannot be saved.
        """
        path = Path(workbook_path)

        try:
            workbook.save(path)
        except (OSError, ValueError) as error:
            raise WorkbookSaveError(
                f"Unable to save workbook to '{path}': {error}"
            ) from error

    def prepare_workbook(self, workbook_path: PathLike) -> tuple[Workbook, Worksheet]:
        """Open a workbook and ensure its active sheet is ready for enrichment.

        The active worksheet must have a Company column. A Location column is
        added when absent. The caller controls where the workbook is saved.
        """
        workbook = self.load_workbook(workbook_path)
        worksheet = self.get_active_worksheet(workbook)
        self.find_company_column(worksheet)
        self.ensure_location_column(worksheet)
        return workbook, worksheet
