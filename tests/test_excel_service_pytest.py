"""Pytest unit tests for Excel workbook service behavior."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from excel_service import (
    CompanyColumnNotFoundError,
    ExcelService,
    WorkbookNotFoundError,
)
from models import Location


@pytest.fixture
def service() -> ExcelService:
    """Return the workbook service under test."""

    return ExcelService()


def test_prepare_and_enrich_workbook_without_modifying_input(
    tmp_path: Path, service: ExcelService
) -> None:
    """Preparation finds headers and output writing is delegated to the service."""

    input_path = tmp_path / "companies.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([" company "])
    worksheet.append(["Example Corp"])
    worksheet.append(["  "])
    workbook.save(input_path)

    prepared_workbook, prepared_sheet = service.prepare_workbook(input_path)
    company_column = service.find_company_column(prepared_sheet)
    location_column = service.ensure_location_column(prepared_sheet)

    assert list(service.iter_company_rows(prepared_sheet, company_column)) == [
        (2, "Example Corp")
    ]
    service.write_location(
        prepared_sheet,
        2,
        location_column,
        Location("Pune", "Maharashtra", "India"),
    )
    output_path = service.output_path(input_path, tmp_path / "output")
    service.save_workbook(prepared_workbook, output_path)

    assert output_path.name == "companies_enriched.xlsx"
    assert load_workbook(output_path).active.cell(row=2, column=2).value == (
        "Pune, Maharashtra, India"
    )
    assert load_workbook(input_path).active.max_column == 1


def test_find_workbooks_is_sorted_and_ignores_temporary_files(
    tmp_path: Path, service: ExcelService
) -> None:
    """Only supported Excel workbooks are discovered in deterministic order."""

    for name in ("z.xlsx", "A.xlsm", "~$locked.xlsx", "notes.txt"):
        (tmp_path / name).touch()

    assert [path.name for path in service.find_workbooks(tmp_path)] == [
        "A.xlsm",
        "z.xlsx",
    ]


def test_find_workbooks_and_company_header_fail_with_clear_errors(
    tmp_path: Path, service: ExcelService
) -> None:
    """Missing workbooks and a missing Company column are rejected."""

    with pytest.raises(WorkbookNotFoundError):
        service.find_workbooks(tmp_path)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Name"])

    with pytest.raises(CompanyColumnNotFoundError):
        service.find_company_column(worksheet)
