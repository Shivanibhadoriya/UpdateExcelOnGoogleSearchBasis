"""Integration-style tests for application orchestration."""

from pathlib import Path
import sys
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from config import Config
from main import _run_enrichment
from models import SearchResult


class ApplicationOrchestrationTests(unittest.TestCase):
    """Verify services are coordinated without embedding their logic in main."""

    def setUp(self) -> None:
        """Create isolated input, output, cache, and log directories."""

        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.input_directory = self.root / "input"
        self.input_directory.mkdir()
        self.output_directory = self.root / "output"
        self.config = Config(
            serp_api_key="test-key",
            input_directory=str(self.input_directory),
            output_directory=str(self.output_directory),
            cache_directory=str(self.root / "cache"),
            log_directory=str(self.root / "logs"),
            request_timeout_seconds=30,
            max_retry_attempts=1,
        )

    def tearDown(self) -> None:
        """Remove the temporary application workspace."""

        self.temporary_directory.cleanup()

    def test_run_enrichment_writes_locations_to_an_output_workbook(self) -> None:
        """An uncached company is searched, cached, and written to output."""

        input_path = self.input_directory / "companies.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Company"])
        worksheet.append(["Example Corp"])
        workbook.save(input_path)

        search_service = Mock()
        search_service.search.return_value = SearchResult(
            company_name="Example Corp",
            knowledge_graph={"India Office": "Pune, Maharashtra, India"},
        )

        with patch("main.SerpApiSearchService", return_value=search_service):
            summary = _run_enrichment(self.config, Mock())

        output_path = self.output_directory / "companies_enriched.xlsx"
        output_worksheet = load_workbook(output_path).active
        self.assertEqual(
            output_worksheet.cell(row=2, column=2).value,
            "Pune, Maharashtra, India",
        )
        self.assertEqual(summary.companies_processed, 1)
        self.assertEqual(summary.locations_found, 1)
        self.assertTrue((self.root / "cache").is_dir())


if __name__ == "__main__":
    unittest.main()
